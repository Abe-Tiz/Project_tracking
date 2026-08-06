# controllers/report_controller.py - Comprehensive Report with Tracking Sheet

from flask import request, make_response
from flask_jwt_extended import get_jwt_identity
from models.project_model import ProjectModel
from models.task_model import TaskModel
from models.report_model import ReportModel
import logging
from datetime import datetime, timedelta
from bson import ObjectId
import json
import math

logger = logging.getLogger(__name__)

class ReportController:
    """Report Controller - Comprehensive Project Report Generator"""
    
    @staticmethod
    def generate_project_report(project_id):
        """Generate a comprehensive project report"""
        try:
            user_id = get_jwt_identity()
            data = request.get_json() or {}
            
            logger.info(f"Generating comprehensive report for project {project_id}")
            
            project = ProjectModel.find_by_id(project_id)
            if not project:
                return {'error': 'Project not found'}, 404
            
            if project.get('owner_id') != user_id and user_id not in project.get('members', []):
                return {'error': 'Access denied'}, 403
            
            try:
                tasks = TaskModel.find_by_project(project_id)
                if tasks is None:
                    tasks = []
            except Exception as e:
                logger.warning(f"Error fetching tasks: {str(e)}")
                tasks = []
            
            try:
                members = ProjectModel.get_members(project_id)
                if members is None:
                    members = []
            except Exception as e:
                logger.warning(f"Error fetching members: {str(e)}")
                members = []
            
            report_data = ReportController._build_comprehensive_report(project, tasks, members, data)
            
            report = {
                'project_id': project_id,
                'project_name': project.get('name', 'Unknown'),
                'generated_by': user_id,
                'generated_by_name': data.get('generated_by_name', 'Unknown'),
                'report_type': data.get('report_type', 'comprehensive'),
                'data': report_data,
                'format': data.get('format', 'json'),
                'status': 'completed',
                'generated_at': datetime.utcnow().isoformat(),
                'date_range': data.get('date_range', {}),
                'company': data.get('company', ''),
                'department': data.get('department', '')
            }
            
            report_id = ReportModel.create(report)
            
            if not report_id:
                return {'error': 'Failed to save report'}, 500
            
            saved_report = ReportModel.find_by_id(report_id)
            return {
                'message': 'Report generated successfully',
                'report_id': report_id,
                'report': ReportModel.to_dict(saved_report) if saved_report else None
            }, 200
            
        except Exception as e:
            logger.error(f"Error generating report: {str(e)}", exc_info=True)
            return {'error': f'Failed to generate report: {str(e)}'}, 500
    
    @staticmethod
    def _build_comprehensive_report(project, tasks, members, params):
        """Build comprehensive international standard report with tracking sheet"""
        try:
            # Initialize report structure
            report = {
                'header': {},
                'executive_summary': {},
                'project_overview': {},
                'tracking_sheet': [],
                'performance_metrics': {},
                'task_analysis': {},
                'team_analysis': {},
                'risk_analysis': {},
                'financial_analysis': {},
                'quality_metrics': {},
                'timeline_analysis': {},
                'recommendations': {},
                'appendix': {}
            }
            
            # Header Information
            report['header'] = {
                'report_title': 'PROJECT STATUS REPORT',
                'project_name': project.get('name', ''),
                'project_id': str(project.get('_id', '')),
                'report_date': datetime.utcnow().strftime('%B %d, %Y'),
                'report_period': params.get('date_range', {}),
                'company': params.get('company', ''),
                'department': params.get('department', ''),
                'report_type': params.get('report_type', 'Comprehensive'),
                'version': '1.0',
                'classification': 'Internal Use'
            }
            
            # Project Overview
            report['project_overview'] = {
                'project_name': project.get('name', ''),
                'project_description': project.get('description', ''),
                'project_status': project.get('status', 'Planning'),
                'project_priority': project.get('priority', 'Medium'),
                'start_date': project.get('start_date', ''),
                'end_date': project.get('end_date', ''),
                'owner': project.get('owner_name', ''),
                'total_members': len(members),
                'total_tasks': len(tasks),
                'created_date': project.get('created_at', datetime.utcnow()).isoformat() if project.get('created_at') else '',
                'last_updated': project.get('updated_at', datetime.utcnow()).isoformat() if project.get('updated_at') else ''
            }
            
            # Tracking Sheet (SoftPOS style)
            report['tracking_sheet'] = ReportController._create_tracking_sheet(tasks, members)
            
            # Executive Summary
            summary = ReportController._calculate_summary_stats(tasks)
            report['executive_summary'] = summary
            
            # Performance Metrics
            report['performance_metrics'] = ReportController._calculate_performance_metrics(tasks, members)
            
            # Task Analysis
            report['task_analysis'] = ReportController._analyze_tasks(tasks)
            
            # Team Analysis
            report['team_analysis'] = ReportController._analyze_team(members, tasks)
            
            # Risk Analysis
            report['risk_analysis'] = ReportController._analyze_risks(tasks, project)
            
            # Financial Analysis
            report['financial_analysis'] = ReportController._analyze_financials(tasks, project)
            
            # Quality Metrics
            report['quality_metrics'] = ReportController._calculate_quality_metrics(tasks)
            
            # Timeline Analysis
            report['timeline_analysis'] = ReportController._analyze_timeline(tasks)
            
            # Recommendations
            report['recommendations'] = ReportController._generate_recommendations(report)
            
            # Appendix
            report['appendix'] = {
                'task_list': ReportController._get_task_list(tasks),
                'member_list': ReportController._get_member_list(members),
                'data_sources': [
                    'Project Management System',
                    'Task Tracking Database',
                    'Team Member Records',
                    'Project Documentation'
                ],
                'methodology': 'Data extracted from project management system and analyzed using standard project management metrics.',
                'glossary': {
                    'EVM': 'Earned Value Management',
                    'CPI': 'Cost Performance Index',
                    'SPI': 'Schedule Performance Index',
                    'KPI': 'Key Performance Indicator',
                    'ROI': 'Return on Investment',
                    'MTTR': 'Mean Time to Resolve'
                }
            }
            
            return report
            
        except Exception as e:
            logger.error(f"Error building comprehensive report: {str(e)}", exc_info=True)
            return {
                'header': {},
                'executive_summary': {'error': str(e)},
                'project_overview': {},
                'tracking_sheet': [],
                'performance_metrics': {},
                'task_analysis': {},
                'team_analysis': {},
                'risk_analysis': {},
                'financial_analysis': {},
                'quality_metrics': {},
                'timeline_analysis': {},
                'recommendations': {},
                'appendix': {}
            }
    
    @staticmethod
    def _create_tracking_sheet(tasks, members):
        """Create SoftPOS style tracking sheet"""
        tracking_sheet = []
        
        # Get member names for lookup
        member_map = {m.get('_id'): m.get('name', 'Unknown') for m in members}
        
        for task in tasks:
            assigned_to_id = task.get('assigned_to', '')
            assigned_to_name = member_map.get(assigned_to_id, task.get('assigned_to_name', 'Unassigned'))
            
            # Determine status with color coding
            status = task.get('status', 'Todo')
            status_display = status
            status_color = {
                'Done': 'completed 100%',
                'In Progress': 'in progress',
                'Review': 'review',
                'Todo': 'pending'
            }.get(status, status)
            
            tracking_sheet.append({
                'requesting_team': task.get('requesting_team', assigned_to_name),
                'request_description': task.get('title', ''),
                'assigned_to': assigned_to_name,
                'request_date': task.get('created_at', '').isoformat() if task.get('created_at') else '',
                'response_date': task.get('due_date', ''),
                'remarks_notes': task.get('description', ''),
                'current_status': status_display,
                'status_color': status_color,
                'priority': task.get('priority', 'Medium'),
                'labels': task.get('labels', [])
            })
        
        return tracking_sheet
    
    @staticmethod
    def _calculate_summary_stats(tasks):
        """Calculate comprehensive summary statistics"""
        total = len(tasks)
        if total == 0:
            return {
                'total_tasks': 0,
                'completed': 0,
                'in_progress': 0,
                'review': 0,
                'todo': 0,
                'completion_rate': 0,
                'on_track': 0,
                'delayed': 0,
                'at_risk': 0,
                'overall_status': 'Not Started'
            }
        
        completed = len([t for t in tasks if t.get('status') == 'Done'])
        in_progress = len([t for t in tasks if t.get('status') == 'In Progress'])
        review = len([t for t in tasks if t.get('status') == 'Review'])
        todo = len([t for t in tasks if t.get('status') == 'Todo'])
        
        completion_rate = round((completed / total) * 100, 2) if total > 0 else 0
        
        # Calculate on-track, delayed, at-risk
        on_track = 0
        delayed = 0
        at_risk = 0
        now = datetime.utcnow()
        
        for task in tasks:
            due_date = task.get('due_date')
            status = task.get('status', '')
            
            if status == 'Done':
                on_track += 1
            elif due_date:
                try:
                    if isinstance(due_date, str):
                        due_date = datetime.fromisoformat(due_date)
                    if now > due_date:
                        delayed += 1
                    elif (due_date - now).days < 7:
                        at_risk += 1
                    else:
                        on_track += 1
                except:
                    on_track += 1
            else:
                on_track += 1
        
        # Determine overall status
        if completion_rate >= 80:
            overall_status = 'On Track'
        elif completion_rate >= 50:
            overall_status = 'Progressing'
        elif completion_rate >= 20:
            overall_status = 'Behind Schedule'
        else:
            overall_status = 'At Risk'
        
        return {
            'total_tasks': total,
            'completed': completed,
            'in_progress': in_progress,
            'review': review,
            'todo': todo,
            'completion_rate': completion_rate,
            'on_track': on_track,
            'delayed': delayed,
            'at_risk': at_risk,
            'overall_status': overall_status
        }
    
    @staticmethod
    def _calculate_performance_metrics(tasks, members):
        """Calculate performance metrics"""
        metrics = {
            'schedule_performance': {},
            'cost_performance': {},
            'productivity': {},
            'quality': {}
        }
        
        total = len(tasks)
        if total == 0:
            return metrics
        
        # Schedule Performance
        completed = len([t for t in tasks if t.get('status') == 'Done'])
        planned_completion = len([t for t in tasks if t.get('due_date')])
        
        spi = (completed / planned_completion) if planned_completion > 0 else 0
        schedule_variance = completed - planned_completion
        
        metrics['schedule_performance'] = {
            'planned_completion': planned_completion,
            'actual_completion': completed,
            'schedule_performance_index': round(spi, 2),
            'schedule_variance': schedule_variance,
            'variance_percentage': round((schedule_variance / planned_completion * 100) if planned_completion > 0 else 0, 2),
            'status': 'Good' if spi >= 0.95 else 'Warning' if spi >= 0.8 else 'Critical'
        }
        
        # Productivity
        total_hours = sum([t.get('estimated_hours', 0) for t in tasks])
        actual_hours = sum([t.get('actual_hours', 0) for t in tasks if t.get('status') == 'Done'])
        
        completed_tasks = [t for t in tasks if t.get('status') == 'Done' and t.get('created_at')]
        velocity = 0
        if completed_tasks:
            first_date = min([t.get('created_at') for t in completed_tasks if t.get('created_at')])
            last_date = max([t.get('created_at') for t in completed_tasks if t.get('created_at')])
            if first_date and last_date:
                days = (last_date - first_date).days if isinstance(last_date, datetime) and isinstance(first_date, datetime) else 30
                weeks = max(days / 7, 1)
                velocity = round(len(completed_tasks) / weeks, 2)
        
        metrics['productivity'] = {
            'total_estimated_hours': total_hours,
            'actual_hours': actual_hours,
            'hours_saved': total_hours - actual_hours if actual_hours > 0 else 0,
            'velocity': velocity,
            'avg_completion_time': round(sum([(t.get('updated_at', t.get('created_at', datetime.utcnow())) - t.get('created_at', datetime.utcnow())).days for t in tasks if t.get('status') == 'Done']), 2) if completed > 0 else 0,
            'efficiency': round((actual_hours / total_hours * 100) if total_hours > 0 else 0, 2)
        }
        
        return metrics
    
    @staticmethod
    def _analyze_tasks(tasks):
        """Analyze tasks in detail"""
        analysis = {
            'by_status': {},
            'by_priority': {},
            'by_label': {},
            'by_assignee': {},
            'age_analysis': {},
            'trends': {}
        }
        
        if not tasks:
            return analysis
        
        # Age analysis
        now = datetime.utcnow()
        age_groups = {
            '0-7': 0,
            '8-14': 0,
            '15-30': 0,
            '30+': 0
        }
        
        for task in tasks:
            created = task.get('created_at')
            if created and isinstance(created, datetime):
                days = (now - created).days
                if days <= 7:
                    age_groups['0-7'] += 1
                elif days <= 14:
                    age_groups['8-14'] += 1
                elif days <= 30:
                    age_groups['15-30'] += 1
                else:
                    age_groups['30+'] += 1
        
        analysis['age_analysis'] = {
            'labels': ['0-7 days', '8-14 days', '15-30 days', '30+ days'],
            'values': [age_groups['0-7'], age_groups['8-14'], age_groups['15-30'], age_groups['30+']],
            'avg_age': round(sum([(now - t.get('created_at')).days for t in tasks if t.get('created_at')]) / len(tasks), 2) if tasks else 0
        }
        
        # Trends (last 30 days)
        thirty_days_ago = now - timedelta(days=30)
        recent_tasks = [t for t in tasks if t.get('created_at') and t.get('created_at') > thirty_days_ago]
        completed_recent = [t for t in recent_tasks if t.get('status') == 'Done']
        
        analysis['trends'] = {
            'tasks_created_last_30_days': len(recent_tasks),
            'tasks_completed_last_30_days': len(completed_recent),
            'completion_trend': 'Increasing' if len(completed_recent) > len(recent_tasks) * 0.3 else 'Stable' if len(completed_recent) > len(recent_tasks) * 0.1 else 'Decreasing',
            'velocity': round(len(completed_recent) / 4, 2) if len(completed_recent) > 0 else 0
        }
        
        return analysis
    
    @staticmethod
    def _analyze_team(members, tasks):
        """Analyze team performance"""
        analysis = {
            'total_members': len(members),
            'active_members': 0,
            'member_performance': [],
            'workload_distribution': {},
            'team_efficiency': 0
        }
        
        if not members:
            return analysis
        
        total_tasks = len(tasks)
        active_count = 0
        
        for member in members:
            member_id = member.get('_id')
            member_tasks = [t for t in tasks if t.get('assigned_to') == member_id]
            completed = len([t for t in member_tasks if t.get('status') == 'Done'])
            
            if member_tasks:
                active_count += 1
            
            analysis['member_performance'].append({
                'name': member.get('name', 'Unknown'),
                'role': member.get('role', 'Team Member'),
                'department': member.get('department', ''),
                'total_tasks': len(member_tasks),
                'completed': completed,
                'completion_rate': round((completed / len(member_tasks) * 100) if len(member_tasks) > 0 else 0, 2),
                'is_external': member.get('is_external', False)
            })
        
        analysis['active_members'] = active_count
        
        # Workload distribution
        workload = [m['total_tasks'] for m in analysis['member_performance']]
        if workload:
            avg_workload = sum(workload) / len(workload) if workload else 0
            max_workload = max(workload) if workload else 0
            min_workload = min(workload) if workload else 0
            
            analysis['workload_distribution'] = {
                'avg_tasks_per_member': round(avg_workload, 2),
                'max_tasks': max_workload,
                'min_tasks': min_workload,
                'balance_score': round((1 - (max_workload - min_workload) / (max_workload + min_workload + 1)) * 100, 2)
            }
        
        # Team efficiency
        completed_tasks = sum([m['completed'] for m in analysis['member_performance']])
        total_assigned = sum([m['total_tasks'] for m in analysis['member_performance']])
        analysis['team_efficiency'] = round((completed_tasks / total_assigned * 100) if total_assigned > 0 else 0, 2)
        
        return analysis
    
    @staticmethod
    def _analyze_risks(tasks, project):
        """Analyze project risks"""
        risks = {
            'high': [],
            'medium': [],
            'low': [],
            'summary': {}
        }
        
        total = len(tasks)
        if total == 0:
            return risks
        
        completed = len([t for t in tasks if t.get('status') == 'Done'])
        in_progress = len([t for t in tasks if t.get('status') == 'In Progress'])
        todo = len([t for t in tasks if t.get('status') == 'Todo'])
        review = len([t for t in tasks if t.get('status') == 'Review'])
        
        # Schedule risk
        if completed / total < 0.3:
            risks['high'].append({
                'category': 'Schedule',
                'description': 'Low completion rate: Only {:.1f}% of tasks completed'.format((completed/total)*100),
                'impact': 'Project may be significantly delayed',
                'mitigation': 'Review priorities and allocate additional resources'
            })
        elif completed / total < 0.6:
            risks['medium'].append({
                'category': 'Schedule',
                'description': 'Moderate completion rate: {:.1f}% of tasks completed'.format((completed/total)*100),
                'impact': 'Project may experience minor delays',
                'mitigation': 'Monitor progress closely and adjust schedule'
            })
        
        # Scope risk
        if todo / total > 0.4:
            risks['high'].append({
                'category': 'Scope',
                'description': 'Large backlog: {} tasks not started ({}%)'.format(todo, round((todo/total)*100, 1)),
                'impact': 'Significant work remains',
                'mitigation': 'Break down tasks and prioritize critical items'
            })
        
        # Resource risk
        if in_progress / total > 0.5:
            risks['medium'].append({
                'category': 'Resources',
                'description': 'Many tasks in progress: {} ({}%)'.format(in_progress, round((in_progress/total)*100, 1)),
                'impact': 'Resource contention possible',
                'mitigation': 'Balance workload across team members'
            })
        
        # Review bottleneck
        if review / total > 0.2:
            risks['medium'].append({
                'category': 'Quality',
                'description': 'Tasks pending review: {} ({}%)'.format(review, round((review/total)*100, 1)),
                'impact': 'Quality assurance may become a bottleneck',
                'mitigation': 'Streamline review process and allocate more reviewers'
            })
        
        risks['summary'] = {
            'total_risks': len(risks['high']) + len(risks['medium']) + len(risks['low']),
            'high_risks': len(risks['high']),
            'medium_risks': len(risks['medium']),
            'low_risks': len(risks['low']),
            'risk_level': 'High' if len(risks['high']) > 2 else 'Medium' if len(risks['high']) > 0 else 'Low'
        }
        
        return risks
    
    @staticmethod
    def _analyze_financials(tasks, project):
        """Analyze financial metrics"""
        financials = {
            'budget': 0,
            'spent': 0,
            'remaining': 0,
            'variance': 0,
            'roi': 0,
            'cost_performance_index': 0,
            'estimated_total_cost': 0,
            'status': 'On Budget'
        }
        
        total_estimated_hours = sum([t.get('estimated_hours', 0) for t in tasks])
        total_actual_hours = sum([t.get('actual_hours', 0) for t in tasks if t.get('status') == 'Done'])
        
        hourly_rate = 50
        estimated_cost = total_estimated_hours * hourly_rate
        actual_cost = total_actual_hours * hourly_rate
        
        financials['budget'] = estimated_cost
        financials['spent'] = actual_cost
        financials['remaining'] = estimated_cost - actual_cost
        financials['variance'] = estimated_cost - actual_cost
        financials['estimated_total_cost'] = estimated_cost
        
        if actual_cost > 0:
            financials['cost_performance_index'] = round(estimated_cost / actual_cost, 2)
        
        completed_tasks = len([t for t in tasks if t.get('status') == 'Done'])
        if completed_tasks > 0:
            value_delivered = completed_tasks * 100
            cost = actual_cost if actual_cost > 0 else 1
            financials['roi'] = round(((value_delivered - cost) / cost) * 100, 2)
        
        if actual_cost > estimated_cost * 1.1:
            financials['status'] = 'Over Budget'
        elif actual_cost > estimated_cost:
            financials['status'] = 'Slightly Over Budget'
        elif actual_cost < estimated_cost * 0.9:
            financials['status'] = 'Under Budget'
        else:
            financials['status'] = 'On Budget'
        
        return financials
    
    @staticmethod
    def _calculate_quality_metrics(tasks):
        """Calculate quality metrics"""
        quality = {
            'defect_rate': 0,
            'review_coverage': 0,
            'test_coverage': 0,
            'rework_percentage': 0,
            'overall_quality_score': 0
        }
        
        total = len(tasks)
        if total == 0:
            return quality
        
        has_review = 0
        has_test = 0
        has_defects = 0
        
        for task in tasks:
            labels = task.get('labels', [])
            if isinstance(labels, str):
                try:
                    labels = json.loads(labels)
                except:
                    labels = []
            
            if 'review' in [l.lower() if isinstance(l, str) else '' for l in labels]:
                has_review += 1
            if 'test' in [l.lower() if isinstance(l, str) else '' for l in labels]:
                has_test += 1
            if 'bug' in [l.lower() if isinstance(l, str) else '' for l in labels] or task.get('priority') == 'High':
                has_defects += 1
        
        quality['review_coverage'] = round((has_review / total) * 100, 2)
        quality['test_coverage'] = round((has_test / total) * 100, 2)
        quality['defect_rate'] = round((has_defects / total) * 100, 2)
        quality['rework_percentage'] = round((has_defects / max(total, 1)) * 100, 2)
        
        quality['overall_quality_score'] = round(
            (quality['review_coverage'] * 0.4 + 
             quality['test_coverage'] * 0.3 + 
             (100 - quality['defect_rate']) * 0.3), 2
        )
        
        return quality
    
    @staticmethod
    def _analyze_timeline(tasks):
        """Analyze project timeline"""
        timeline = {
            'start_date': None,
            'end_date': None,
            'duration_days': 0,
            'on_time_completion': 0,
            'delayed_completion': 0,
            'average_task_duration': 0,
            'critical_path': []
        }
        
        if not tasks:
            return timeline
        
        start_dates = [t.get('created_at') for t in tasks if t.get('created_at')]
        end_dates = [t.get('updated_at') for t in tasks if t.get('status') == 'Done' and t.get('updated_at')]
        
        if start_dates:
            timeline['start_date'] = min(start_dates).isoformat() if start_dates else None
        if end_dates:
            timeline['end_date'] = max(end_dates).isoformat() if end_dates else None
        
        if timeline['start_date'] and timeline['end_date']:
            start = datetime.fromisoformat(timeline['start_date'])
            end = datetime.fromisoformat(timeline['end_date'])
            timeline['duration_days'] = (end - start).days
        
        completed_tasks = [t for t in tasks if t.get('status') == 'Done']
        on_time = 0
        delayed = 0
        
        for task in completed_tasks:
            due_date = task.get('due_date')
            completed_at = task.get('updated_at')
            if due_date and completed_at:
                if isinstance(due_date, str):
                    try:
                        due_date = datetime.fromisoformat(due_date)
                    except:
                        due_date = None
                if due_date and completed_at <= due_date:
                    on_time += 1
                else:
                    delayed += 1
        
        total_completed = len(completed_tasks)
        timeline['on_time_completion'] = round((on_time / total_completed * 100) if total_completed > 0 else 0, 2)
        timeline['delayed_completion'] = round((delayed / total_completed * 100) if total_completed > 0 else 0, 2)
        
        durations = []
        for task in completed_tasks:
            created = task.get('created_at')
            updated = task.get('updated_at')
            if created and updated:
                days = (updated - created).days
                durations.append(days)
        
        timeline['average_task_duration'] = round(sum(durations) / len(durations), 2) if durations else 0
        
        sorted_tasks = sorted(completed_tasks, key=lambda t: (t.get('updated_at', datetime.utcnow()) - t.get('created_at', datetime.utcnow())).days if t.get('updated_at') and t.get('created_at') else 0, reverse=True)
        critical_path = []
        for task in sorted_tasks[:5]:
            if task.get('updated_at') and task.get('created_at'):
                days = (task['updated_at'] - task['created_at']).days
                critical_path.append({
                    'title': task.get('title', ''),
                    'duration_days': days,
                    'status': task.get('status', '')
                })
        
        timeline['critical_path'] = critical_path
        
        return timeline
    
    @staticmethod
    def _generate_recommendations(report):
        """Generate actionable recommendations"""
        recommendations = {
            'high_priority': [],
            'medium_priority': [],
            'low_priority': [],
            'summary': ''
        }
        
        summary = report.get('executive_summary', {})
        performance = report.get('performance_metrics', {})
        risks = report.get('risk_analysis', {})
        quality = report.get('quality_metrics', {})
        team = report.get('team_analysis', {})
        
        # High priority recommendations
        if summary.get('completion_rate', 100) < 50:
            recommendations['high_priority'].append({
                'area': 'Progress',
                'recommendation': 'Project completion rate is below 50%. Immediate action required to accelerate progress.',
                'action_items': [
                    'Re-evaluate project priorities',
                    'Allocate additional resources to critical tasks',
                    'Remove blockers and dependencies'
                ]
            })
        
        if risks.get('summary', {}).get('high_risks', 0) > 2:
            recommendations['high_priority'].append({
                'area': 'Risk Management',
                'recommendation': 'Multiple high-severity risks identified. Urgent mitigation required.',
                'action_items': [
                    'Implement risk mitigation plans immediately',
                    'Schedule risk review meetings',
                    'Escalate critical risks to stakeholders'
                ]
            })
        
        if summary.get('delayed', 0) > summary.get('total_tasks', 1) * 0.3:
            recommendations['high_priority'].append({
                'area': 'Schedule',
                'recommendation': 'Significant number of delayed tasks. Schedule recovery needed.',
                'action_items': [
                    'Review and adjust project timeline',
                    'Identify and address bottlenecks',
                    'Consider fast-tracking or crashing'
                ]
            })
        
        # Medium priority recommendations
        if team.get('workload_distribution', {}).get('balance_score', 100) < 60:
            recommendations['medium_priority'].append({
                'area': 'Team Workload',
                'recommendation': 'Uneven workload distribution among team members.',
                'action_items': [
                    'Reassign tasks to balance workload',
                    'Provide support to overloaded members',
                    'Cross-train team members'
                ]
            })
        
        if quality.get('overall_quality_score', 100) < 70:
            recommendations['medium_priority'].append({
                'area': 'Quality',
                'recommendation': 'Quality metrics below threshold. Quality improvement needed.',
                'action_items': [
                    'Strengthen review process',
                    'Increase testing coverage',
                    'Address defect patterns'
                ]
            })
        
        if performance.get('schedule_performance', {}).get('status', 'Good') == 'Critical':
            recommendations['medium_priority'].append({
                'area': 'Performance',
                'recommendation': 'Critical schedule performance index. Performance improvement required.',
                'action_items': [
                    'Analyze root causes of delays',
                    'Implement performance improvement plan',
                    'Monitor KPIs closely'
                ]
            })
        
        # Low priority recommendations
        if summary.get('completion_rate', 100) < 80:
            recommendations['low_priority'].append({
                'area': 'Continuous Improvement',
                'recommendation': 'Consider process improvements for better efficiency.',
                'action_items': [
                    'Conduct retrospective sessions',
                    'Identify best practices',
                    'Implement process improvements'
                ]
            })
        
        if recommendations['high_priority']:
            recommendations['summary'] = 'High priority actions required. Focus on progress acceleration and risk mitigation.'
        elif recommendations['medium_priority']:
            recommendations['summary'] = 'Medium priority improvements recommended. Focus on team optimization and quality.'
        elif recommendations['low_priority']:
            recommendations['summary'] = 'Minor improvements suggested. Monitor progress and implement gradually.'
        else:
            recommendations['summary'] = 'Project is on track. Continue monitoring and maintain current momentum.'
        
        return recommendations
    
    @staticmethod
    def _get_task_list(tasks):
        """Get formatted task list for appendix"""
        task_list = []
        for task in tasks[:50]:
            task_list.append({
                'id': str(task.get('_id', '')),
                'title': task.get('title', ''),
                'status': task.get('status', ''),
                'priority': task.get('priority', ''),
                'assigned_to': task.get('assigned_to_name', ''),
                'due_date': task.get('due_date', ''),
                'estimated_hours': task.get('estimated_hours', 0),
                'created_at': task.get('created_at', '').isoformat() if task.get('created_at') else ''
            })
        return task_list
    
    @staticmethod
    def _get_member_list(members):
        """Get formatted member list for appendix"""
        member_list = []
        for member in members:
            member_list.append({
                'id': member.get('_id', ''),
                'name': member.get('name', ''),
                'email': member.get('email', ''),
                'role': member.get('role', ''),
                'department': member.get('department', ''),
                'is_external': member.get('is_external', False)
            })
        return member_list
    
    @staticmethod
    def export_report(report_id, format_type='pdf'):
        """Export a report in specified format"""
        try:
            user_id = get_jwt_identity()
            
            report = ReportModel.find_by_id(report_id)
            if not report:
                return {'error': 'Report not found'}, 404
            
            project = ProjectModel.find_by_id(report.get('project_id'))
            if project:
                if project.get('owner_id') != user_id and user_id not in project.get('members', []):
                    return {'error': 'Access denied'}, 403
            
            report_data = report.get('data', {})
            project_name = report.get('project_name', 'project')
            
            logger.info(f"Exporting comprehensive report as {format_type.upper()}")
            
            if format_type == 'pdf':
                return ReportController._export_comprehensive_pdf(report_data, project_name, report)
            elif format_type == 'excel':
                return ReportController._export_comprehensive_excel(report_data, project_name, report)
            elif format_type == 'csv':
                return ReportController._export_csv(report_data, project_name, report)
            elif format_type == 'json':
                return ReportController._export_json(report_data, project_name, report)
            else:
                return {'error': f'Unsupported format: {format_type}'}, 400
                
        except Exception as e:
            logger.error(f"Error exporting report: {str(e)}", exc_info=True)
            return {'error': f'Failed to export report: {str(e)}'}, 500

    @staticmethod
    def _export_json(report_data, project_name, report):
        """Export as JSON"""
        from flask import jsonify
        
        response_data = {
            'report': report_data,
            'metadata': {
                'project_name': project_name,
                'generated_at': report.get('generated_at'),
                'report_type': report.get('report_type')
            }
        }
        
        response = make_response(jsonify(response_data))
        response.headers['Content-Disposition'] = f'attachment; filename={project_name}_report.json'
        response.headers['Content-Type'] = 'application/json'
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    @staticmethod
    def _export_csv(report_data, project_name, report):
        """Export as CSV"""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['PROJECT REPORT - COMPREHENSIVE'])
        writer.writerow([f'Project: {project_name}'])
        writer.writerow([f'Generated: {report.get("generated_at")}'])
        writer.writerow([])
        
        # Executive Summary
        summary = report_data.get('executive_summary', {})
        writer.writerow(['EXECUTIVE SUMMARY'])
        writer.writerow(['Metric', 'Value'])
        for key, value in summary.items():
            writer.writerow([key, value])
        writer.writerow([])
        
        # Tracking Sheet
        tracking = report_data.get('tracking_sheet', [])
        if tracking:
            writer.writerow(['TRACKING SHEET'])
            writer.writerow(['Requesting Team', 'Request Description', 'Assigned To', 'Request Date', 'Response Date', 'Remarks/Notes', 'Current Status'])
            for item in tracking:
                writer.writerow([
                    item.get('requesting_team', ''),
                    item.get('request_description', ''),
                    item.get('assigned_to', ''),
                    item.get('request_date', ''),
                    item.get('response_date', ''),
                    item.get('remarks_notes', ''),
                    item.get('current_status', '')
                ])
            writer.writerow([])
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={project_name}_report.csv'
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    # @staticmethod
    # def _export_comprehensive_excel(report_data, project_name, report):
    #     """Export comprehensive report as Excel with Tracking Sheet"""
    #     try:
    #         import openpyxl
    #         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    #         from openpyxl.utils import get_column_letter
    #         from io import BytesIO
            
    #         logger.info(f"Generating comprehensive Excel report for {project_name}")
            
    #         wb = openpyxl.Workbook()
            
    #         # Styles
    #         title_font = Font(bold=True, size=16, color='1B4F72')
    #         header_font = Font(bold=True, size=11, color='FFFFFF')
    #         header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    #         sub_header_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    #         completed_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    #         in_progress_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    #         review_fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
    #         todo_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
    #         border = Border(
    #             left=Side(style='thin', color='CCCCCC'),
    #             right=Side(style='thin', color='CCCCCC'),
    #             top=Side(style='thin', color='CCCCCC'),
    #             bottom=Side(style='thin', color='CCCCCC')
    #         )
            
    #         # ============ SHEET 1: TRACKING SHEET (SoftPOS Style) ============
    #         ws1 = wb.active
    #         ws1.title = 'Tracking Sheet'
            
    #         # Set column widths for tracking sheet
    #         ws1.column_dimensions['A'].width = 25  # Requesting Team
    #         ws1.column_dimensions['B'].width = 45  # Request Description
    #         ws1.column_dimensions['C'].width = 20  # Assigned To
    #         ws1.column_dimensions['D'].width = 15  # Request Date
    #         ws1.column_dimensions['E'].width = 15  # Response Date
    #         ws1.column_dimensions['F'].width = 35  # Remarks/Notes
    #         ws1.column_dimensions['G'].width = 18  # Current Status
    #         ws1.column_dimensions['H'].width = 12  # Priority
    #         ws1.column_dimensions['I'].width = 25  # Labels
            
    #         # Title
    #         ws1.merge_cells('A1:I1')
    #         cell = ws1['A1']
    #         cell.value = f'{project_name} - TRACKING SHEET'
    #         cell.font = Font(bold=True, size=18, color='1B4F72')
    #         cell.alignment = Alignment(horizontal='center')
            
    #         # Subtitle
    #         ws1.merge_cells('A2:I2')
    #         cell = ws1['A2']
    #         cell.value = f'Generated: {report.get("generated_at", datetime.utcnow().isoformat())}'
    #         cell.font = Font(size=10, color='5D6D7E')
    #         cell.alignment = Alignment(horizontal='center')
            
    #         # Headers
    #         headers = ['Requesting Team', 'Request Description', 'Assigned To', 'Request Date', 'Response Date', 'Remarks/Notes', 'Current Status', 'Priority', 'Labels']
    #         for col, header in enumerate(headers, 1):
    #             cell = ws1.cell(row=4, column=col, value=header)
    #             cell.font = header_font
    #             cell.fill = header_fill
    #             cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    #             cell.border = border
            
    #         # Data - Tracking Sheet
    #         tracking = report_data.get('tracking_sheet', [])
    #         row = 5
    #         for idx, item in enumerate(tracking, 1):
    #             ws1.cell(row=row, column=1, value=item.get('requesting_team', 'N/A'))
    #             ws1.cell(row=row, column=2, value=item.get('request_description', ''))
    #             ws1.cell(row=row, column=3, value=item.get('assigned_to', 'Unassigned'))
    #             ws1.cell(row=row, column=4, value=item.get('request_date', ''))
    #             ws1.cell(row=row, column=5, value=item.get('response_date', ''))
    #             ws1.cell(row=row, column=6, value=item.get('remarks_notes', ''))
                
    #             # Status with color
    #             status = item.get('current_status', 'Todo')
    #             status_cell = ws1.cell(row=row, column=7, value=status)
    #             if status == 'Done' or status == 'completed 100%':
    #                 status_cell.font = Font(color='2E7D32')
    #                 status_cell.fill = completed_fill
    #             elif status == 'In Progress' or status == 'in progress':
    #                 status_cell.font = Font(color='E65100')
    #                 status_cell.fill = in_progress_fill
    #             elif status == 'Review' or status == 'review':
    #                 status_cell.font = Font(color='6A1B9A')
    #                 status_cell.fill = review_fill
    #             else:
    #                 status_cell.fill = todo_fill
                
    #             ws1.cell(row=row, column=8, value=item.get('priority', 'Medium'))
    #             ws1.cell(row=row, column=9, value=', '.join(item.get('labels', [])))
                
    #             # Apply border to all cells
    #             for col in range(1, 10):
    #                 ws1.cell(row=row, column=col).border = border
    #                 ws1.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                
    #             row += 1
            
    #         # ============ SHEET 2: EXECUTIVE SUMMARY ============
    #         ws2 = wb.create_sheet('Executive Summary')
    #         ws2.column_dimensions['A'].width = 30
    #         ws2.column_dimensions['B'].width = 25
            
    #         ws2.merge_cells('A1:B1')
    #         cell = ws2['A1']
    #         cell.value = 'EXECUTIVE SUMMARY'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         summary = report_data.get('executive_summary', {})
    #         row = 3
    #         for key, value in summary.items():
    #             ws2[f'A{row}'] = key.replace('_', ' ').title()
    #             ws2[f'A{row}'].font = Font(bold=True)
    #             ws2[f'B{row}'] = value
    #             ws2[f'A{row}'].border = border
    #             ws2[f'B{row}'].border = border
    #             row += 1
            
    #         # ============ SHEET 3: PROJECT OVERVIEW ============
    #         ws3 = wb.create_sheet('Project Overview')
    #         ws3.column_dimensions['A'].width = 25
    #         ws3.column_dimensions['B'].width = 30
            
    #         ws3.merge_cells('A1:B1')
    #         cell = ws3['A1']
    #         cell.value = 'PROJECT OVERVIEW'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         overview = report_data.get('project_overview', {})
    #         row = 3
    #         for key, value in overview.items():
    #             ws3[f'A{row}'] = key.replace('_', ' ').title()
    #             ws3[f'A{row}'].font = Font(bold=True)
    #             ws3[f'B{row}'] = value
    #             ws3[f'A{row}'].border = border
    #             ws3[f'B{row}'].border = border
    #             row += 1
            
    #         # ============ SHEET 4: PERFORMANCE METRICS ============
    #         ws4 = wb.create_sheet('Performance Metrics')
    #         ws4.column_dimensions['A'].width = 30
    #         ws4.column_dimensions['B'].width = 20
    #         ws4.column_dimensions['C'].width = 15
            
    #         ws4.merge_cells('A1:C1')
    #         cell = ws4['A1']
    #         cell.value = 'PERFORMANCE METRICS'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         performance = report_data.get('performance_metrics', {})
    #         row = 3
    #         for category, metrics in performance.items():
    #             ws4[f'A{row}'] = category.replace('_', ' ').title()
    #             ws4[f'A{row}'].font = Font(bold=True, size=12)
    #             ws4.merge_cells(f'A{row}:C{row}')
    #             row += 1
    #             for key, value in metrics.items():
    #                 if not isinstance(value, dict):
    #                     ws4[f'A{row}'] = f'  {key.replace("_", " ").title()}'
    #                     ws4[f'B{row}'] = value
    #                     ws4[f'A{row}'].border = border
    #                     ws4[f'B{row}'].border = border
    #                     row += 1
    #             row += 1
            
    #         # ============ SHEET 5: TASK ANALYSIS ============
    #         ws5 = wb.create_sheet('Task Analysis')
    #         ws5.column_dimensions['A'].width = 20
    #         ws5.column_dimensions['B'].width = 15
    #         ws5.column_dimensions['C'].width = 15
            
    #         ws5.merge_cells('A1:C1')
    #         cell = ws5['A1']
    #         cell.value = 'TASK ANALYSIS'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         # Status breakdown
    #         ws5['A3'] = 'STATUS BREAKDOWN'
    #         ws5['A3'].font = Font(bold=True, size=12)
    #         ws5.merge_cells('A3:C3')
            
    #         task_analysis = report_data.get('task_analysis', {})
    #         statuses = ['Todo', 'In Progress', 'Review', 'Done']
    #         row = 4
    #         for status in statuses:
    #             count = len([t for t in report_data.get('appendix', {}).get('task_list', []) if t.get('status') == status])
    #             ws5[f'A{row}'] = status
    #             ws5[f'B{row}'] = count
    #             ws5[f'A{row}'].border = border
    #             ws5[f'B{row}'].border = border
    #             row += 1
            
    #         # Priority breakdown
    #         ws5[f'A{row+2}'] = 'PRIORITY BREAKDOWN'
    #         ws5[f'A{row+2}'].font = Font(bold=True, size=12)
    #         ws5.merge_cells(f'A{row+2}:C{row+2}')
            
    #         priorities = ['High', 'Medium', 'Low']
    #         row += 3
    #         for priority in priorities:
    #             count = len([t for t in report_data.get('appendix', {}).get('task_list', []) if t.get('priority') == priority])
    #             ws5[f'A{row}'] = priority
    #             ws5[f'B{row}'] = count
    #             ws5[f'A{row}'].border = border
    #             ws5[f'B{row}'].border = border
    #             row += 1
            
    #         # Age analysis
    #         ws5[f'A{row+2}'] = 'AGE ANALYSIS'
    #         ws5[f'A{row+2}'].font = Font(bold=True, size=12)
    #         ws5.merge_cells(f'A{row+2}:C{row+2}')
            
    #         age_data = task_analysis.get('age_analysis', {})
    #         row += 3
    #         for label, value in zip(age_data.get('labels', []), age_data.get('values', [])):
    #             ws5[f'A{row}'] = label
    #             ws5[f'B{row}'] = value
    #             ws5[f'A{row}'].border = border
    #             ws5[f'B{row}'].border = border
    #             row += 1
            
    #         # ============ SHEET 6: TEAM ANALYSIS ============
    #         ws6 = wb.create_sheet('Team Analysis')
    #         ws6.column_dimensions['A'].width = 25
    #         ws6.column_dimensions['B'].width = 20
    #         ws6.column_dimensions['C'].width = 15
    #         ws6.column_dimensions['D'].width = 15
    #         ws6.column_dimensions['E'].width = 15
            
    #         ws6.merge_cells('A1:E1')
    #         cell = ws6['A1']
    #         cell.value = 'TEAM PERFORMANCE'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         headers = ['Name', 'Role', 'Total Tasks', 'Completed', 'Rate']
    #         for col, header in enumerate(headers, 1):
    #             cell = ws6.cell(row=3, column=col, value=header)
    #             cell.font = header_font
    #             cell.fill = header_fill
    #             cell.alignment = Alignment(horizontal='center')
    #             cell.border = border
            
    #         team = report_data.get('team_analysis', {})
    #         members = team.get('member_performance', [])
    #         row = 4
    #         for member in members:
    #             ws6.cell(row=row, column=1, value=member.get('name', ''))
    #             ws6.cell(row=row, column=2, value=member.get('role', ''))
    #             ws6.cell(row=row, column=3, value=member.get('total_tasks', 0))
    #             ws6.cell(row=row, column=4, value=member.get('completed', 0))
    #             ws6.cell(row=row, column=5, value=f"{member.get('completion_rate', 0)}%")
    #             for col in range(1, 6):
    #                 ws6.cell(row=row, column=col).border = border
    #             row += 1
            
    #         # Workload distribution
    #         row += 2
    #         workload = team.get('workload_distribution', {})
    #         workload_headers = ['Metric', 'Value']
    #         for col, header in enumerate(workload_headers, 1):
    #             cell = ws6.cell(row=row, column=col, value=header)
    #             cell.font = header_font
    #             cell.fill = header_fill
    #             cell.alignment = Alignment(horizontal='center')
    #             cell.border = border
            
    #         row += 1
    #         for key, value in workload.items():
    #             ws6.cell(row=row, column=1, value=key.replace('_', ' ').title())
    #             ws6.cell(row=row, column=2, value=value)
    #             for col in range(1, 3):
    #                 ws6.cell(row=row, column=col).border = border
    #             row += 1
            
    #         # ============ SHEET 7: RISK ANALYSIS ============
    #         ws7 = wb.create_sheet('Risk Analysis')
    #         ws7.column_dimensions['A'].width = 15
    #         ws7.column_dimensions['B'].width = 30
    #         ws7.column_dimensions['C'].width = 30
    #         ws7.column_dimensions['D'].width = 30
            
    #         ws7.merge_cells('A1:D1')
    #         cell = ws7['A1']
    #         cell.value = 'RISK ANALYSIS'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         risk_headers = ['Level', 'Category', 'Description', 'Mitigation']
    #         for col, header in enumerate(risk_headers, 1):
    #             cell = ws7.cell(row=3, column=col, value=header)
    #             cell.font = header_font
    #             cell.fill = header_fill
    #             cell.alignment = Alignment(horizontal='center')
    #             cell.border = border
            
    #         risks = report_data.get('risk_analysis', {})
    #         row = 4
    #         for level in ['high', 'medium', 'low']:
    #             for risk in risks.get(level, []):
    #                 ws7.cell(row=row, column=1, value=level.upper())
    #                 ws7.cell(row=row, column=2, value=risk.get('category', ''))
    #                 ws7.cell(row=row, column=3, value=risk.get('description', ''))
    #                 ws7.cell(row=row, column=4, value=risk.get('mitigation', ''))
    #                 for col in range(1, 5):
    #                     ws7.cell(row=row, column=col).border = border
    #                 row += 1
            
    #         # ============ SHEET 8: FINANCIAL ANALYSIS ============
    #         ws8 = wb.create_sheet('Financial Analysis')
    #         ws8.column_dimensions['A'].width = 25
    #         ws8.column_dimensions['B'].width = 20
            
    #         ws8.merge_cells('A1:B1')
    #         cell = ws8['A1']
    #         cell.value = 'FINANCIAL ANALYSIS'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         financials = report_data.get('financial_analysis', {})
    #         row = 3
    #         for key, value in financials.items():
    #             ws8[f'A{row}'] = key.replace('_', ' ').title()
    #             ws8[f'A{row}'].font = Font(bold=True)
    #             ws8[f'B{row}'] = value
    #             ws8[f'A{row}'].border = border
    #             ws8[f'B{row}'].border = border
    #             row += 1
            
    #         # ============ SHEET 9: RECOMMENDATIONS ============
    #         ws9 = wb.create_sheet('Recommendations')
    #         ws9.column_dimensions['A'].width = 15
    #         ws9.column_dimensions['B'].width = 25
    #         ws9.column_dimensions['C'].width = 25
    #         ws9.column_dimensions['D'].width = 35
            
    #         ws9.merge_cells('A1:D1')
    #         cell = ws9['A1']
    #         cell.value = 'RECOMMENDATIONS'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         rec_headers = ['Priority', 'Area', 'Recommendation', 'Action Items']
    #         for col, header in enumerate(rec_headers, 1):
    #             cell = ws9.cell(row=3, column=col, value=header)
    #             cell.font = header_font
    #             cell.fill = header_fill
    #             cell.alignment = Alignment(horizontal='center')
    #             cell.border = border
            
    #         recommendations = report_data.get('recommendations', {})
    #         row = 4
    #         for priority in ['high_priority', 'medium_priority', 'low_priority']:
    #             for rec in recommendations.get(priority, []):
    #                 priority_label = priority.replace('_', ' ').title()
    #                 ws9.cell(row=row, column=1, value=priority_label)
    #                 ws9.cell(row=row, column=2, value=rec.get('area', ''))
    #                 ws9.cell(row=row, column=3, value=rec.get('recommendation', ''))
    #                 action_items = ', '.join(rec.get('action_items', []))
    #                 ws9.cell(row=row, column=4, value=action_items)
    #                 for col in range(1, 5):
    #                     ws9.cell(row=row, column=col).border = border
    #                 row += 1
            
    #         # ============ SHEET 10: QUALITY METRICS ============
    #         ws10 = wb.create_sheet('Quality Metrics')
    #         ws10.column_dimensions['A'].width = 25
    #         ws10.column_dimensions['B'].width = 20
            
    #         ws10.merge_cells('A1:B1')
    #         cell = ws10['A1']
    #         cell.value = 'QUALITY METRICS'
    #         cell.font = title_font
    #         cell.alignment = Alignment(horizontal='center')
            
    #         quality = report_data.get('quality_metrics', {})
    #         row = 3
    #         for key, value in quality.items():
    #             ws10[f'A{row}'] = key.replace('_', ' ').title()
    #             ws10[f'A{row}'].font = Font(bold=True)
    #             ws10[f'B{row}'] = value
    #             ws10[f'A{row}'].border = border
    #             ws10[f'B{row}'].border = border
    #             row += 1
            
    #         # Auto-adjust column widths
    #         for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8, ws9, ws10]:
    #             for column in ws.columns:
    #                 max_length = 0
    #                 column_letter = get_column_letter(column[0].column)
    #                 for cell in column:
    #                     try:
    #                         if cell.value and len(str(cell.value)) > max_length:
    #                             max_length = len(str(cell.value))
    #                     except:
    #                         pass
    #                 adjusted_width = min(max_length + 3, 50)
    #                 ws.column_dimensions[column_letter].width = adjusted_width
            
    #         output = BytesIO()
    #         wb.save(output)
    #         output.seek(0)
            
    #         logger.info(f"Comprehensive Excel report generated for {project_name}")
            
    #         response = make_response(output.getvalue())
    #         response.headers['Content-Disposition'] = f'attachment; filename={project_name}_comprehensive_report.xlsx'
    #         response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #         response.headers['Access-Control-Allow-Origin'] = '*'
    #         response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    #         response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
    #         return response
            
    #     except ImportError as e:
    #         logger.error(f"openpyxl not installed: {str(e)}")
    #         return ReportController._export_csv(report_data, project_name, report)
    #     except Exception as e:
    #         logger.error(f"Error exporting comprehensive Excel: {str(e)}", exc_info=True)
    #         return ReportController._export_csv(report_data, project_name, report)

    
    
    @staticmethod
    def _export_comprehensive_excel(report_data, project_name, report):
        """Export comprehensive report as Excel with Tracking Sheet"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            logger.info(f"Generating comprehensive Excel report for {project_name}")
            
            wb = openpyxl.Workbook()
            
            # Styles
            title_font = Font(bold=True, size=16, color='1B4F72')
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
            sub_header_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
            completed_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            in_progress_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
            review_fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
            todo_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # ============ SHEET 1: TRACKING SHEET (SoftPOS Style) ============
            ws1 = wb.active
            ws1.title = 'Tracking Sheet'
            
            # Set column widths for tracking sheet
            ws1.column_dimensions['A'].width = 25  # Requesting Team
            ws1.column_dimensions['B'].width = 45  # Request Description
            ws1.column_dimensions['C'].width = 20  # Assigned To
            ws1.column_dimensions['D'].width = 15  # Request Date
            ws1.column_dimensions['E'].width = 15  # Response Date
            ws1.column_dimensions['F'].width = 35  # Remarks/Notes
            ws1.column_dimensions['G'].width = 18  # Current Status
            ws1.column_dimensions['H'].width = 12  # Priority
            ws1.column_dimensions['I'].width = 25  # Labels
            
            # Title
            ws1.merge_cells('A1:I1')
            cell = ws1['A1']
            cell.value = f'{project_name} - TRACKING SHEET'
            cell.font = Font(bold=True, size=18, color='1B4F72')
            cell.alignment = Alignment(horizontal='center')
            
            # Subtitle
            ws1.merge_cells('A2:I2')
            cell = ws1['A2']
            cell.value = f'Generated: {report.get("generated_at", datetime.utcnow().isoformat())}'
            cell.font = Font(size=10, color='5D6D7E')
            cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ['Requesting Team', 'Request Description', 'Assigned To', 'Request Date', 'Response Date', 'Remarks/Notes', 'Current Status', 'Priority', 'Labels']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Data - Tracking Sheet
            tracking = report_data.get('tracking_sheet', [])
            row = 5
            for idx, item in enumerate(tracking, 1):
                # Get labels and filter out None values
                labels = item.get('labels', [])
                if labels is None:
                    labels = []
                # Filter out None values and convert to string
                labels_str = ', '.join([str(l) for l in labels if l is not None])
                
                ws1.cell(row=row, column=1, value=item.get('requesting_team', 'N/A') or 'N/A')
                ws1.cell(row=row, column=2, value=item.get('request_description', '') or '')
                ws1.cell(row=row, column=3, value=item.get('assigned_to', 'Unassigned') or 'Unassigned')
                ws1.cell(row=row, column=4, value=item.get('request_date', '') or '')
                ws1.cell(row=row, column=5, value=item.get('response_date', '') or '')
                ws1.cell(row=row, column=6, value=item.get('remarks_notes', '') or '')
                
                # Status with color
                status = item.get('current_status', 'Todo') or 'Todo'
                status_cell = ws1.cell(row=row, column=7, value=status)
                if status == 'Done' or status == 'completed 100%':
                    status_cell.font = Font(color='2E7D32')
                    status_cell.fill = completed_fill
                elif status == 'In Progress' or status == 'in progress':
                    status_cell.font = Font(color='E65100')
                    status_cell.fill = in_progress_fill
                elif status == 'Review' or status == 'review':
                    status_cell.font = Font(color='6A1B9A')
                    status_cell.fill = review_fill
                else:
                    status_cell.fill = todo_fill
                
                ws1.cell(row=row, column=8, value=item.get('priority', 'Medium') or 'Medium')
                ws1.cell(row=row, column=9, value=labels_str)
                
                # Apply border to all cells
                for col in range(1, 10):
                    ws1.cell(row=row, column=col).border = border
                    ws1.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                
                row += 1
            
            # ============ SHEET 2: EXECUTIVE SUMMARY ============
            ws2 = wb.create_sheet('Executive Summary')
            ws2.column_dimensions['A'].width = 30
            ws2.column_dimensions['B'].width = 25
            
            ws2.merge_cells('A1:B1')
            cell = ws2['A1']
            cell.value = 'EXECUTIVE SUMMARY'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            summary = report_data.get('executive_summary', {})
            row = 3
            for key, value in summary.items():
                if value is not None:
                    ws2[f'A{row}'] = str(key).replace('_', ' ').title()
                    ws2[f'A{row}'].font = Font(bold=True)
                    ws2[f'B{row}'] = str(value)
                    ws2[f'A{row}'].border = border
                    ws2[f'B{row}'].border = border
                    row += 1
            
            # ============ SHEET 3: PROJECT OVERVIEW ============
            ws3 = wb.create_sheet('Project Overview')
            ws3.column_dimensions['A'].width = 25
            ws3.column_dimensions['B'].width = 30
            
            ws3.merge_cells('A1:B1')
            cell = ws3['A1']
            cell.value = 'PROJECT OVERVIEW'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            overview = report_data.get('project_overview', {})
            row = 3
            for key, value in overview.items():
                if value is not None:
                    ws3[f'A{row}'] = str(key).replace('_', ' ').title()
                    ws3[f'A{row}'].font = Font(bold=True)
                    ws3[f'B{row}'] = str(value)
                    ws3[f'A{row}'].border = border
                    ws3[f'B{row}'].border = border
                    row += 1
            
            # ============ SHEET 4: PERFORMANCE METRICS ============
            ws4 = wb.create_sheet('Performance Metrics')
            ws4.column_dimensions['A'].width = 30
            ws4.column_dimensions['B'].width = 20
            ws4.column_dimensions['C'].width = 15
            
            ws4.merge_cells('A1:C1')
            cell = ws4['A1']
            cell.value = 'PERFORMANCE METRICS'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            performance = report_data.get('performance_metrics', {})
            row = 3
            for category, metrics in performance.items():
                if metrics:
                    ws4[f'A{row}'] = str(category).replace('_', ' ').title()
                    ws4[f'A{row}'].font = Font(bold=True, size=12)
                    ws4.merge_cells(f'A{row}:C{row}')
                    row += 1
                    for key, value in metrics.items():
                        if not isinstance(value, dict) and value is not None:
                            ws4[f'A{row}'] = f'  {str(key).replace("_", " ").title()}'
                            ws4[f'B{row}'] = str(value)
                            ws4[f'A{row}'].border = border
                            ws4[f'B{row}'].border = border
                            row += 1
                    row += 1
            
            # ============ SHEET 5: TASK ANALYSIS ============
            ws5 = wb.create_sheet('Task Analysis')
            ws5.column_dimensions['A'].width = 20
            ws5.column_dimensions['B'].width = 15
            ws5.column_dimensions['C'].width = 15
            
            ws5.merge_cells('A1:C1')
            cell = ws5['A1']
            cell.value = 'TASK ANALYSIS'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            # Status breakdown
            ws5['A3'] = 'STATUS BREAKDOWN'
            ws5['A3'].font = Font(bold=True, size=12)
            ws5.merge_cells('A3:C3')
            
            task_analysis = report_data.get('task_analysis', {})
            statuses = ['Todo', 'In Progress', 'Review', 'Done']
            row = 4
            for status in statuses:
                count = len([t for t in report_data.get('appendix', {}).get('task_list', []) if t.get('status') == status])
                ws5[f'A{row}'] = status
                ws5[f'B{row}'] = count
                ws5[f'A{row}'].border = border
                ws5[f'B{row}'].border = border
                row += 1
            
            # Priority breakdown
            ws5[f'A{row+2}'] = 'PRIORITY BREAKDOWN'
            ws5[f'A{row+2}'].font = Font(bold=True, size=12)
            ws5.merge_cells(f'A{row+2}:C{row+2}')
            
            priorities = ['High', 'Medium', 'Low']
            row += 3
            for priority in priorities:
                count = len([t for t in report_data.get('appendix', {}).get('task_list', []) if t.get('priority') == priority])
                ws5[f'A{row}'] = priority
                ws5[f'B{row}'] = count
                ws5[f'A{row}'].border = border
                ws5[f'B{row}'].border = border
                row += 1
            
            # Age analysis
            ws5[f'A{row+2}'] = 'AGE ANALYSIS'
            ws5[f'A{row+2}'].font = Font(bold=True, size=12)
            ws5.merge_cells(f'A{row+2}:C{row+2}')
            
            age_data = task_analysis.get('age_analysis', {})
            row += 3
            labels = age_data.get('labels', [])
            values = age_data.get('values', [])
            for label, value in zip(labels, values):
                if label is not None and value is not None:
                    ws5[f'A{row}'] = str(label)
                    ws5[f'B{row}'] = value
                    ws5[f'A{row}'].border = border
                    ws5[f'B{row}'].border = border
                    row += 1
            
            # ============ SHEET 6: TEAM ANALYSIS ============
            ws6 = wb.create_sheet('Team Analysis')
            ws6.column_dimensions['A'].width = 25
            ws6.column_dimensions['B'].width = 20
            ws6.column_dimensions['C'].width = 15
            ws6.column_dimensions['D'].width = 15
            ws6.column_dimensions['E'].width = 15
            
            ws6.merge_cells('A1:E1')
            cell = ws6['A1']
            cell.value = 'TEAM PERFORMANCE'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            headers = ['Name', 'Role', 'Total Tasks', 'Completed', 'Rate']
            for col, header in enumerate(headers, 1):
                cell = ws6.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            team = report_data.get('team_analysis', {})
            members = team.get('member_performance', [])
            row = 4
            for member in members:
                ws6.cell(row=row, column=1, value=member.get('name', '') or 'Unknown')
                ws6.cell(row=row, column=2, value=member.get('role', '') or 'Team Member')
                ws6.cell(row=row, column=3, value=member.get('total_tasks', 0) or 0)
                ws6.cell(row=row, column=4, value=member.get('completed', 0) or 0)
                rate = member.get('completion_rate', 0) or 0
                ws6.cell(row=row, column=5, value=f"{rate}%")
                for col in range(1, 6):
                    ws6.cell(row=row, column=col).border = border
                row += 1
            
            # Workload distribution
            row += 2
            workload = team.get('workload_distribution', {})
            if workload:
                workload_headers = ['Metric', 'Value']
                for col, header in enumerate(workload_headers, 1):
                    cell = ws6.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                row += 1
                for key, value in workload.items():
                    if value is not None:
                        ws6.cell(row=row, column=1, value=str(key).replace('_', ' ').title())
                        ws6.cell(row=row, column=2, value=str(value))
                        for col in range(1, 3):
                            ws6.cell(row=row, column=col).border = border
                        row += 1
            
            # ============ SHEET 7: RISK ANALYSIS ============
            ws7 = wb.create_sheet('Risk Analysis')
            ws7.column_dimensions['A'].width = 15
            ws7.column_dimensions['B'].width = 30
            ws7.column_dimensions['C'].width = 30
            ws7.column_dimensions['D'].width = 30
            
            ws7.merge_cells('A1:D1')
            cell = ws7['A1']
            cell.value = 'RISK ANALYSIS'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            risk_headers = ['Level', 'Category', 'Description', 'Mitigation']
            for col, header in enumerate(risk_headers, 1):
                cell = ws7.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            risks = report_data.get('risk_analysis', {})
            row = 4
            for level in ['high', 'medium', 'low']:
                for risk in risks.get(level, []):
                    ws7.cell(row=row, column=1, value=level.upper())
                    ws7.cell(row=row, column=2, value=risk.get('category', '') or '')
                    ws7.cell(row=row, column=3, value=risk.get('description', '') or '')
                    ws7.cell(row=row, column=4, value=risk.get('mitigation', '') or '')
                    for col in range(1, 5):
                        ws7.cell(row=row, column=col).border = border
                    row += 1
            
            # ============ SHEET 8: FINANCIAL ANALYSIS ============
            ws8 = wb.create_sheet('Financial Analysis')
            ws8.column_dimensions['A'].width = 25
            ws8.column_dimensions['B'].width = 20
            
            ws8.merge_cells('A1:B1')
            cell = ws8['A1']
            cell.value = 'FINANCIAL ANALYSIS'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            financials = report_data.get('financial_analysis', {})
            row = 3
            for key, value in financials.items():
                if value is not None:
                    ws8[f'A{row}'] = str(key).replace('_', ' ').title()
                    ws8[f'A{row}'].font = Font(bold=True)
                    ws8[f'B{row}'] = str(value)
                    ws8[f'A{row}'].border = border
                    ws8[f'B{row}'].border = border
                    row += 1
            
            # ============ SHEET 9: RECOMMENDATIONS ============
            ws9 = wb.create_sheet('Recommendations')
            ws9.column_dimensions['A'].width = 15
            ws9.column_dimensions['B'].width = 25
            ws9.column_dimensions['C'].width = 25
            ws9.column_dimensions['D'].width = 35
            
            ws9.merge_cells('A1:D1')
            cell = ws9['A1']
            cell.value = 'RECOMMENDATIONS'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            rec_headers = ['Priority', 'Area', 'Recommendation', 'Action Items']
            for col, header in enumerate(rec_headers, 1):
                cell = ws9.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            recommendations = report_data.get('recommendations', {})
            row = 4
            for priority in ['high_priority', 'medium_priority', 'low_priority']:
                for rec in recommendations.get(priority, []):
                    priority_label = str(priority).replace('_', ' ').title()
                    ws9.cell(row=row, column=1, value=priority_label)
                    ws9.cell(row=row, column=2, value=rec.get('area', '') or '')
                    ws9.cell(row=row, column=3, value=rec.get('recommendation', '') or '')
                    action_items = rec.get('action_items', [])
                    if action_items:
                        # Filter out None values and convert to string
                        action_items_str = ', '.join([str(item) for item in action_items if item is not None])
                    else:
                        action_items_str = ''
                    ws9.cell(row=row, column=4, value=action_items_str)
                    for col in range(1, 5):
                        ws9.cell(row=row, column=col).border = border
                    row += 1
            
            # ============ SHEET 10: QUALITY METRICS ============
            ws10 = wb.create_sheet('Quality Metrics')
            ws10.column_dimensions['A'].width = 25
            ws10.column_dimensions['B'].width = 20
            
            ws10.merge_cells('A1:B1')
            cell = ws10['A1']
            cell.value = 'QUALITY METRICS'
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            
            quality = report_data.get('quality_metrics', {})
            row = 3
            for key, value in quality.items():
                if value is not None:
                    ws10[f'A{row}'] = str(key).replace('_', ' ').title()
                    ws10[f'A{row}'].font = Font(bold=True)
                    ws10[f'B{row}'] = str(value)
                    ws10[f'A{row}'].border = border
                    ws10[f'B{row}'].border = border
                    row += 1
            
            # Auto-adjust column widths
            for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8, ws9, ws10]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 3, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"Comprehensive Excel report generated for {project_name}")
            
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename={project_name}_comprehensive_report.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
            response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
            
        except ImportError as e:
            logger.error(f"openpyxl not installed: {str(e)}")
            return ReportController._export_csv(report_data, project_name, report)
        except Exception as e:
            logger.error(f"Error exporting comprehensive Excel: {str(e)}", exc_info=True)
            return ReportController._export_csv(report_data, project_name, report)

            
        

    @staticmethod
    def _export_comprehensive_pdf(report_data, project_name, report):
        """Export comprehensive report as PDF with Tracking Sheet"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            from io import BytesIO
            import datetime
            
            logger.info(f"Generating comprehensive PDF report for {project_name}")
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=A4, 
                rightMargin=72, 
                leftMargin=72, 
                topMargin=72, 
                bottomMargin=72
            )
            
            styles = getSampleStyleSheet()
            story = []
            
            # ============ CUSTOM STYLES ============
            cover_title_style = ParagraphStyle(
                'CoverTitle',
                parent=styles['Heading1'],
                fontSize=28,
                textColor=colors.HexColor('#1B4F72'),
                alignment=TA_CENTER,
                spaceAfter=30,
                fontName='Helvetica-Bold'
            )
            
            cover_subtitle_style = ParagraphStyle(
                'CoverSubtitle',
                parent=styles['Heading2'],
                fontSize=18,
                textColor=colors.HexColor('#2E4053'),
                alignment=TA_CENTER,
                spaceAfter=20,
                fontName='Helvetica'
            )
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1B4F72'),
                alignment=TA_CENTER,
                spaceAfter=30,
                fontName='Helvetica-Bold'
            )
            
            section_title_style = ParagraphStyle(
                'SectionTitle',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#1B4F72'),
                spaceAfter=12,
                spaceBefore=20,
                fontName='Helvetica-Bold'
            )
            
            sub_section_style = ParagraphStyle(
                'SubSection',
                parent=styles['Heading3'],
                fontSize=12,
                textColor=colors.HexColor('#2E4053'),
                spaceAfter=8,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            # ============ COVER PAGE ============
            header = report_data.get('header', {})
            story.append(Paragraph('PROJECT STATUS REPORT', cover_title_style))
            story.append(Spacer(1, 10))
            story.append(Paragraph(header.get('project_name', ''), cover_subtitle_style))
            story.append(Spacer(1, 20))
            story.append(Paragraph(f'Report Date: {header.get("report_date", "")}', styles['Normal']))
            story.append(Paragraph(f'Report Type: {header.get("report_type", "Comprehensive")}', styles['Normal']))
            story.append(Paragraph(f'Classification: {header.get("classification", "Internal Use")}', styles['Normal']))
            story.append(Paragraph(f'Version: {header.get("version", "1.0")}', styles['Normal']))
            story.append(PageBreak())
            
            # ============ EXECUTIVE SUMMARY ============
            story.append(Paragraph('EXECUTIVE SUMMARY', title_style))
            story.append(Spacer(1, 10))
            
            summary = report_data.get('executive_summary', {})
            summary_data = []
            for key, value in summary.items():
                if isinstance(value, (int, float, str)):
                    summary_data.append([
                        key.replace('_', ' ').title(),
                        str(value)
                    ])
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2.5*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6EAF8')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ]))
            story.append(summary_table)
            story.append(PageBreak())
            
            # ============ PROJECT OVERVIEW ============
            story.append(Paragraph('PROJECT OVERVIEW', title_style))
            story.append(Spacer(1, 10))
            
            overview = report_data.get('project_overview', {})
            overview_data = []
            for key, value in overview.items():
                overview_data.append([
                    key.replace('_', ' ').title(),
                    str(value)
                ])
            
            overview_table = Table(overview_data, colWidths=[2.5*inch, 3*inch])
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6EAF8')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(overview_table)
            story.append(PageBreak())
            
            # ============ TRACKING SHEET ============
            story.append(Paragraph('TRACKING SHEET', title_style))
            story.append(Spacer(1, 10))
            
            tracking = report_data.get('tracking_sheet', [])
            if tracking:
                # Limit to 20 items for PDF readability
                display_tracking = tracking[:20]
                tracking_data = [['#', 'Requesting Team', 'Request Description', 'Assigned To', 'Status']]
                for idx, item in enumerate(display_tracking, 1):
                    status_color = ''
                    if 'completed' in item.get('current_status', '').lower():
                        status_color = '<font color="green">'
                    elif 'progress' in item.get('current_status', '').lower():
                        status_color = '<font color="orange">'
                    elif 'review' in item.get('current_status', '').lower():
                        status_color = '<font color="purple">'
                    else:
                        status_color = '<font color="gray">'
                    
                    tracking_data.append([
                        str(idx),
                        item.get('requesting_team', '')[:20],
                        item.get('request_description', '')[:40] + ('...' if len(item.get('request_description', '')) > 40 else ''),
                        item.get('assigned_to', ''),
                        f"{status_color}{item.get('current_status', '')}</font>"
                    ])
                
                tracking_table = Table(tracking_data, colWidths=[0.3*inch, 1.2*inch, 2.2*inch, 1*inch, 1*inch])
                tracking_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(tracking_table)
                
                if len(tracking) > 20:
                    story.append(Paragraph(f'<i>Showing 20 of {len(tracking)} tracking entries. Full list available in Excel export.</i>', styles['Italic']))
            
            story.append(PageBreak())
            
            # ============ PERFORMANCE METRICS ============
            story.append(Paragraph('PERFORMANCE METRICS', title_style))
            story.append(Spacer(1, 10))
            
            performance = report_data.get('performance_metrics', {})
            for category, metrics in performance.items():
                story.append(Paragraph(category.replace('_', ' ').title(), sub_section_style))
                
                metric_data = []
                for key, value in metrics.items():
                    if not isinstance(value, dict):
                        metric_data.append([
                            key.replace('_', ' ').title(),
                            str(value)
                        ])
                
                if metric_data:
                    metric_table = Table(metric_data, colWidths=[2.5*inch, 2.5*inch])
                    metric_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6EAF8')),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    story.append(metric_table)
                    story.append(Spacer(1, 10))
            
            story.append(PageBreak())
            
            # ============ TEAM ANALYSIS ============
            story.append(Paragraph('TEAM PERFORMANCE', title_style))
            story.append(Spacer(1, 10))
            
            team = report_data.get('team_analysis', {})
            members = team.get('member_performance', [])
            
            if members:
                member_data = [['Name', 'Role', 'Tasks', 'Completed', 'Rate']]
                for member in members:
                    member_data.append([
                        member.get('name', ''),
                        member.get('role', ''),
                        str(member.get('total_tasks', 0)),
                        str(member.get('completed', 0)),
                        f"{member.get('completion_rate', 0)}%"
                    ])
                
                member_table = Table(member_data, colWidths=[1.5*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.8*inch])
                member_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(member_table)
                story.append(Spacer(1, 15))
                
                # Workload distribution
                workload = team.get('workload_distribution', {})
                story.append(Paragraph('Workload Distribution', sub_section_style))
                workload_data = [
                    ['Avg Tasks/Member', str(workload.get('avg_tasks_per_member', 0))],
                    ['Max Tasks', str(workload.get('max_tasks', 0))],
                    ['Min Tasks', str(workload.get('min_tasks', 0))],
                    ['Balance Score', f"{workload.get('balance_score', 0)}%"]
                ]
                workload_table = Table(workload_data, colWidths=[2.5*inch, 2.5*inch])
                workload_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6EAF8')),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ]))
                story.append(workload_table)
                story.append(Spacer(1, 10))
                story.append(Paragraph(f"Team Efficiency: {team.get('team_efficiency', 0)}%", styles['Normal']))
            
            story.append(PageBreak())
            
            # ============ RISK ANALYSIS ============
            story.append(Paragraph('RISK ANALYSIS', title_style))
            story.append(Spacer(1, 10))
            
            risks = report_data.get('risk_analysis', {})
            risk_summary = risks.get('summary', {})
            
            risk_summary_data = [
                ['Total Risks', str(risk_summary.get('total_risks', 0))],
                ['High Risks', str(risk_summary.get('high_risks', 0))],
                ['Medium Risks', str(risk_summary.get('medium_risks', 0))],
                ['Low Risks', str(risk_summary.get('low_risks', 0))],
                ['Risk Level', risk_summary.get('risk_level', 'Low')]
            ]
            
            risk_summary_table = Table(risk_summary_data, colWidths=[2.5*inch, 2.5*inch])
            risk_summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6EAF8')),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(risk_summary_table)
            story.append(Spacer(1, 15))
            
            # List risks
            for level in ['high', 'medium', 'low']:
                level_risks = risks.get(level, [])
                if level_risks:
                    story.append(Paragraph(f'{level.upper()} Risks', sub_section_style))
                    for risk in level_risks:
                        story.append(Paragraph(f'<b>{risk.get("category", "")}</b>: {risk.get("description", "")}', styles['Normal']))
                        story.append(Paragraph(f'<i>Mitigation: {risk.get("mitigation", "")}</i>', styles['Italic']))
                        story.append(Spacer(1, 5))
            
            story.append(PageBreak())
            
            # ============ RECOMMENDATIONS ============
            story.append(Paragraph('RECOMMENDATIONS', title_style))
            story.append(Spacer(1, 10))
            
            recommendations = report_data.get('recommendations', {})
            for priority in ['high_priority', 'medium_priority', 'low_priority']:
                recs = recommendations.get(priority, [])
                if recs:
                    story.append(Paragraph(priority.replace('_', ' ').title(), sub_section_style))
                    for rec in recs:
                        story.append(Paragraph(f'<b>{rec.get("area", "")}</b>', styles['Normal']))
                        story.append(Paragraph(rec.get('recommendation', ''), styles['Normal']))
                        if rec.get('action_items'):
                            story.append(Paragraph('<i>Action Items:</i>', styles['Italic']))
                            for item in rec.get('action_items', []):
                                story.append(Paragraph(f'• {item}', styles['Normal']))
                        story.append(Spacer(1, 5))
            
            story.append(PageBreak())
            
            # ============ QUALITY METRICS ============
            story.append(Paragraph('QUALITY METRICS', title_style))
            story.append(Spacer(1, 10))
            
            quality = report_data.get('quality_metrics', {})
            quality_data = []
            for key, value in quality.items():
                quality_data.append([
                    key.replace('_', ' ').title(),
                    str(value)
                ])
            
            quality_table = Table(quality_data, colWidths=[2.5*inch, 2.5*inch])
            quality_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6EAF8')),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(quality_table)
            
            story.append(PageBreak())
            
            # ============ APPENDIX ============
            story.append(Paragraph('APPENDIX', title_style))
            story.append(Spacer(1, 10))
            
            # Task list
            tasks = report_data.get('appendix', {}).get('task_list', [])
            if tasks:
                story.append(Paragraph('Task List', sub_section_style))
                task_data = [['#', 'Title', 'Status', 'Priority', 'Assigned To']]
                for idx, task in enumerate(tasks[:20], 1):
                    task_data.append([
                        str(idx),
                        task.get('title', '')[:30] + ('...' if len(task.get('title', '')) > 30 else ''),
                        task.get('status', ''),
                        task.get('priority', ''),
                        task.get('assigned_to', '')
                    ])
                
                task_table = Table(task_data, colWidths=[0.4*inch, 2.5*inch, 0.8*inch, 0.8*inch, 1.5*inch])
                task_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(task_table)
                
                if len(tasks) > 20:
                    story.append(Paragraph(f'<i>Showing 20 of {len(tasks)} tasks. Full list available in Excel export.</i>', styles['Italic']))
            
            # ============ FOOTER ============
            story.append(Spacer(1, 30))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#95A5A6'),
                alignment=TA_CENTER
            )
            story.append(Paragraph(f'Report ID: {report.get("_id", "")} • Generated on {report.get("generated_at", "")}', footer_style))
            
            doc.build(story)
            buffer.seek(0)
            
            logger.info(f"Comprehensive PDF report generated for {project_name}")
            
            response = make_response(buffer.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename={project_name}_comprehensive_report.pdf'
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
            response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
            
        except ImportError as e:
            logger.error(f"reportlab not installed: {str(e)}")
            return ReportController._export_json(report_data, project_name, report)
        except Exception as e:
            logger.error(f"Error exporting comprehensive PDF: {str(e)}", exc_info=True)
            return ReportController._export_json(report_data, project_name, report)
    
    @staticmethod
    def get_report(report_id):
        """Get a specific report by ID"""
        try:
            user_id = get_jwt_identity()
            
            report = ReportModel.find_by_id(report_id)
            if not report:
                return {'error': 'Report not found'}, 404
            
            project = ProjectModel.find_by_id(report.get('project_id'))
            if project:
                if project.get('owner_id') != user_id and user_id not in project.get('members', []):
                    return {'error': 'Access denied'}, 403
            
            return ReportModel.to_dict(report), 200
            
        except Exception as e:
            logger.error(f"Error getting report: {str(e)}", exc_info=True)
            return {'error': f'Failed to get report: {str(e)}'}, 500
    
    @staticmethod
    def get_project_reports(project_id):
        """Get all reports for a project"""
        try:
            user_id = get_jwt_identity()
            
            project = ProjectModel.find_by_id(project_id)
            if not project:
                return {'error': 'Project not found'}, 404
            
            if project.get('owner_id') != user_id and user_id not in project.get('members', []):
                return {'error': 'Access denied'}, 403
            
            page = request.args.get('page', 1, type=int)
            per_page = min(request.args.get('per_page', 20, type=int), 100)
            skip = (page - 1) * per_page
            
            reports = ReportModel.find_by_project(project_id, skip, per_page)
            total = len(reports)
            
            return {
                'reports': ReportModel.to_list(reports),
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': total,
                    'pages': (total + per_page - 1) // per_page if per_page > 0 else 0
                }
            }, 200
            
        except Exception as e:
            logger.error(f"Error getting project reports: {str(e)}", exc_info=True)
            return {'error': f'Failed to get reports: {str(e)}'}, 500
    
    @staticmethod
    def delete_report(report_id):
        """Delete a report"""
        try:
            user_id = get_jwt_identity()
            
            report = ReportModel.find_by_id(report_id)
            if not report:
                return {'error': 'Report not found'}, 404
            
            project = ProjectModel.find_by_id(report.get('project_id'))
            if project and project.get('owner_id') != user_id:
                return {'error': 'Only project owner can delete reports'}, 403
            
            success = ReportModel.delete(report_id)
            
            if success:
                return {'message': 'Report deleted successfully'}, 200
            else:
                return {'error': 'Failed to delete report'}, 500
                
        except Exception as e:
            logger.error(f"Error deleting report: {str(e)}", exc_info=True)
            return {'error': f'Failed to delete report: {str(e)}'}, 500